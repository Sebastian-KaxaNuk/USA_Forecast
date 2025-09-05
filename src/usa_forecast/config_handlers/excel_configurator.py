"""
Loads and returns a Configuration entity from an Excel file

Functions
---------
init:
    Loads and returns Configuration
"""

import types
import typing

from src.usa_forecast.entities.configuration import Configuration
from src.usa_forecast.config_handlers.configurator_interface import ConfiguratorInterface

from src.usa_forecast.exceptions import (
    ConfigurationError,
    ConfigurationHandlerError
)

import logging
import pathlib
import sys

import openpyxl
import openpyxl.cell
import openpyxl.utils.exceptions
import openpyxl.worksheet.cell_range
import openpyxl.worksheet.worksheet
import pandas as pd

logger = logging.getLogger('myAppLogger')

class ExcelConfigurator(ConfiguratorInterface):

    SHEET_KEY_VALUES = types.MappingProxyType({
        'General': (
            'start_date',
            'end_date',
            'fmp_api_key',
            'stay_update',
            'summary_mode',
            'summary_frequency',
            'summary_start_date',
            'summary_end_date',
        ),
    })

    SHEET_COLUMNS = types.MappingProxyType({
        'Tickers': (
            'tickers',
        ),
        'Windows_Shift': (
            'window_shift',
        )
    })

    def __init__(
        self,
        file_path: str,
        # logger_format: str = "[%(levelname)s] %(message)s",
        ):
        """
        Initialize configuration, input handlers and output handlers based on a configuration Excel file.

        Parameters
        ----------
        file_path
            The path to the Excel configuration file
        logger_format
            The format for the logger messages. will be injected to logging.basicConfig()
        """
        # not using logging.basicConfig as we need to close it, without affecting any existing root logger
        # logger = logging.getLogger(__name__)
        # set the logging level to info by default, as we haven't loaded the file configuration yet
        # logger.setLevel(logging.INFO)
        # handler = logging.StreamHandler()
        # handler.setFormatter(
        #     logging.Formatter(logger_format)
        # )
        # logger.addHandler(handler)

        try:
            workbook = self._load_file(file_path)
            sheet_key_values = self._extract_workbook_key_values_by_schema(
                workbook,
                self.SHEET_KEY_VALUES,
            )
            sheet_columns = self._extract_workbook_columns_by_schema(
                workbook,
                self.SHEET_COLUMNS,
            )

            raw_window_shift = sheet_columns["Windows_Shift"]["window_shift"]

            window_shift_converted = tuple(
                int(x) for x in raw_window_shift if pd.notna(x)
            )

            self._configuration = Configuration(
                start_date=sheet_key_values['General']['start_date'].date(),
                end_date=sheet_key_values['General']['end_date'].date(),
                fmp_api_key=sheet_key_values['General']['fmp_api_key'],
                stay_update=sheet_key_values['General']['stay_update'],
                summary_mode=sheet_key_values['General']['summary_mode'],
                summary_frequency=sheet_key_values['General']['summary_frequency'],
                summary_start_date=sheet_key_values['General']['summary_start_date'].date(),
                summary_end_date=sheet_key_values['General']['summary_end_date'].date(),
                tickers=sheet_columns['Tickers']['tickers'],
                window_shift=window_shift_converted
            )

            # remove this logger
            logger.handlers.clear()
        except (
            ConfigurationError,
            ConfigurationHandlerError
        ) as error:
            msg = f"An error was encountered when parsing your configuration file: {error!s}"
            logging.getLogger(__name__).critical(msg)
            sys.exit()

    def get_configuration(self) -> Configuration:
        return self._configuration

    @staticmethod
    def _extract_cell_value(cell: openpyxl.cell.cell.Cell) -> str | None:
        """
        Extract the value of an openpyxl cell as a string.

        Parameters
        ----------
        cell
            The cell object to extract the value from.

        Returns
        -------
        The stripped value of the cell. If the cell is empty returns None.
        """
        if (
            cell is not None
            and cell.value is not None
        ):

            return str(cell.value).strip()

        else:

            return None

    @classmethod
    def _extract_column_values(
        cls,
        column: tuple[openpyxl.cell.cell.Cell, ...]
    ):
        """
        Extract the values of an openpyxl column as a string list.

        Parameters
        ----------
        column : openpyxl.worksheet.cell_range.CellRange
            The column object to extract the values from.

        Returns
        -------
        list[str]
            The values as strings.
        """
        values = filter(
            None,
            (
                cls._extract_cell_value(i)
                for i in column
            )
        )

        return list(values)

    @classmethod
    def _extract_workbook_columns_by_schema(
        cls,
        workbook: openpyxl.workbook.workbook.Workbook,
        schema: typing.Mapping[str, tuple[str, ...]],
        header_row: int=1
    ) -> dict[str, dict[str, tuple[typing.Any, ...]]]:
        """
        Extract the values of particular columns specified in the schema from a workbook's sheets.

        Parameters
        ----------
        workbook
            The workbook to search
        schema
            The schema indicating the required sheets and the columns to be extracted from them
        header_row
            The number of the row that will include the column headings

        Returns
        -------
        The extracted values, in the same arrangement as the schema

        Raises
        ------
        ConfigurationHandlerError
        """
        missing_sheets = set(schema.keys()).difference(
            set(workbook.sheetnames)
        )
        if len(missing_sheets) > 0:
            msg = " ".join([
                "The following sheets are missing from the Configuration file:",
                ", ".join(missing_sheets)
            ])

            raise ConfigurationHandlerError(msg)

        sheets = {}
        missing_sheet_columns = {}
        for sheet_name in schema:
            sheets[sheet_name] = {}
            missing_columns = []
            for column_name in schema[sheet_name]:
                column = cls._find_sheet_column_by_row_value(
                    workbook[sheet_name],
                    header_row,
                    column_name
                )
                if column is None:
                    missing_columns.append(column_name)
                else:
                    sheets[sheet_name][column_name] = cls._extract_column_values(
                        workbook[sheet_name][column][header_row:]
                    )
            if len(missing_columns) > 0:
                missing_sheet_columns[sheet_name] = missing_columns

        if len(missing_sheet_columns) > 0:
            msg = " ".join([
                "The following sheet columns are missing from the Configuration file:",
                "; ".join([
                    f"{sheet}: " + ", ".join(fields)
                    for sheet, fields in missing_sheet_columns.items()
                ])
            ])

            raise ConfigurationHandlerError(msg)

        return sheets

    @classmethod
    def _extract_workbook_key_values_by_schema(
        cls,
        workbook: openpyxl.workbook.workbook.Workbook,
        schema: typing.Mapping[str, tuple[str, ...]],
        key_column: str='A',
    ) -> dict[str, dict[str, typing.Any]]:
        """
        Extract the values of particular key values specified in the schema from a workbook's sheets.

        Parameters
        ----------
        workbook
            The workbook to search
        schema
            The schema indicating the required sheets and the key values to be extracted from them
        key_column
            The letter identifier of the column to search for the key

        Returns
        -------
        The extracted values, in the same arrangement as the schema

        Raises
        ------
        ConfigurationHandlerError
        """
        value_column = cls._increment_column_identifier(key_column)
        missing_sheets = set(schema.keys()).difference(
            set(workbook.sheetnames)
        )
        if len(missing_sheets) > 0:
            msg = " ".join([
                "The following sheets are missing from the Configuration file:",
                ", ".join(missing_sheets)
            ])

            raise ConfigurationHandlerError(msg)

        sheets = {}
        missing_sheet_fields = {}
        for sheet_name in schema:
            sheets[sheet_name] = {}
            missing_fields = []
            for field_name in schema[sheet_name]:
                row = cls._find_sheet_row_by_column_value(
                    workbook[sheet_name],
                    key_column,
                    field_name
                )
                if row is None:
                    missing_fields.append(field_name)
                else:
                    sheets[sheet_name][field_name] = workbook[sheet_name][f'{value_column}{row}'].value
            if len(missing_fields) > 0:
                missing_sheet_fields[sheet_name] = missing_fields

        if len(missing_sheet_fields) > 0:
            msg = " ".join([
                "The following sheet params are missing from the Configuration file:",
                "; ".join([
                    f"{sheet}: " + ", ".join(fields)
                    for sheet, fields in missing_sheet_fields.items()
                ])
            ])

            raise ConfigurationHandlerError(msg)

        return sheets

    @staticmethod
    def _find_sheet_column_by_row_value(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        search_value: str
    ) -> int | None:
        """
        Find the column location of the search_value in the row.

        Parameters
        ----------
        sheet
            The sheet in which to search

        row
            The number of the row where to search

        search_value
            the value to be searched for in the row

        Returns
        -------
        The letter identifier of the column where the value was found, or None if not found
        """
        found_column = None

        for column_number in range(1, sheet.max_column + 1):
            column = chr(
                ord('A') + column_number - 1
            )
            cell_name = f"{column}{row}"
            if (
                sheet[cell_name].value is not None
                and sheet[cell_name].value.strip() == search_value
            ):
                found_column = column
                break

        return found_column

    @staticmethod
    def _find_sheet_row_by_column_value(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        column: str,
        search_value: str
    ) -> int | None:
        """
        Find the row location of the search_value in the column.

        Parameters
        ----------
        sheet
            The sheet in which to search
        column
            The letter identofoer of the column where to search
        search_value
            the value to be searched for in the column

        Returns
        -------
        The number of the row where the value was found, or None if not found
        """
        found_row = None

        for row in range(1, sheet.max_row + 1):
            cell_name = f"{column}{row}"
            if (
                sheet[cell_name].value is not None
                and sheet[cell_name].value.strip() == search_value
            ):
                found_row = row
                break

        return found_row

    @classmethod
    def _increment_column_identifier(
        cls,
        column_identifier: str
    ) -> str:
        """
        Get the letter identifier of the next column after column_identifier.

        Uses Excel's column identifier scheme of successive uppercase letters. AA follows Z, and so on...

        Parameters
        ----------
        column_identifier
            The column identifier to increment

        Returns
        -------
        The next column letter identifier
        """
        if column_identifier == 'Z':

            return 'AA'

        elif column_identifier[-1] == 'Z':

            return cls._increment_column_identifier(column_identifier[:-1]) + 'A'

        else:
            next_character = chr(
                ord(column_identifier[-1])
                + 1
            )

            return column_identifier[:-1] + next_character

    @staticmethod
    def _load_file(file_path: str) -> openpyxl.Workbook:
        """
        Load an Excel file.

        Parameters
        ----------
        file_path
            The path of the Excel file to load

        Returns
        -------
        The openpyxl Workbook corresponding to the file

        Raises
        ------
        ConfigurationHandlerError
        """
        if not pathlib.Path(file_path).is_file():
            msg = f"Configuration file not found in path: {file_path}"

            raise ConfigurationHandlerError(msg)

        try:
            workbook = openpyxl.load_workbook(file_path)
        except openpyxl.utils.exceptions.InvalidFileException as error:
            msg = f"Invalid Configuration file in path: {file_path}"

            raise ConfigurationHandlerError(msg) from error

        return workbook
